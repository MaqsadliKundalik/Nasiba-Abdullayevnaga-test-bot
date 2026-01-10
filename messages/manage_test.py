from aiogram import Router, F
from aiogram.types import Message, FSInputFile
from utils.models import User, Tests, UserAnswers
from tortoise.exceptions import DoesNotExist
from config import ADMIN
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from collections import defaultdict
import os

router = Router()

def check_test_keys_format(test_key: str) -> bool:
    pattern = r'^(?:\d+[a-zA-Z])+$'
    return bool(re.fullmatch(pattern, test_key))

@router.message(F.chat.id == ADMIN,F.text.startswith('new '))
async def manage_test(message: Message):
    test_keys = message.text.split()[1]
    user = await User.get(tg_id=message.from_user.id)
    if not check_test_keys_format(test_keys):
        await message.answer("Test kaliti noto'g'ri formatda. Iltimos, to'g'ri formatda kiriting (masalan: 1a2b3c...).")
        return

    test = await Tests.create(user=user, test_keys=test_keys)
    await message.answer(f"Yangi test yaratildi!\n\nTest kodi: `{test.id}`", parse_mode="MARKDOWN")

@router.message(F.chat.id == ADMIN, F.text.startswith('stop '))
async def stop_test(message: Message):
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Buyruq noto'g'ri. Namuna: /stop_test <test_id>")
        return
    test_code = parts[1].strip()
    try:
        test_id = int(test_code)
    except ValueError:
        await message.answer("Test ID butun son bo'lishi kerak.")
        return

    try:
        user = await User.get(tg_id=message.from_user.id)
    except DoesNotExist:
        await message.answer("Foydalanuvchi topilmadi.")
        return

    test = await Tests.filter(id=test_id, user=user).first()
    if not test:
        await message.answer("Sizda bunday test mavjud emas.")
        return
    if test.status == 'stopped':
        await message.answer("Bu test allaqachon to'xtatilgan.")
        return
    test.status = 'stopped'
    await test.save()
    top_answers = await (UserAnswers
                         .filter(test=test)
                         .select_related('user')
                         .order_by('-score', 'created_at')
                         .limit(30))

    if not top_answers:
        await message.answer("Bu testda hali natijalar yo'q.")
        return

    lines = ["Natijalar:\n"]
    for idx, ans in enumerate(top_answers, 1):
        name = ans.user.name or f"ID:{ans.user.id}"
        time_spent = (ans.created_at - test.created_at).total_seconds()
        match idx:
            case 1:
                idx = "🥇"
            case 2:
                idx = "🥈"
            case 3:
                idx = "🥉"

        lines.append(f"{idx}. {name} - {ans.score} ta ({round(time_spent / 60, 1)} min)") 
    msg = "\n".join(lines)
    # jami nechta odam qatnashdi
    msg += f"\n\nJami qatnashganlar soni: {await UserAnswers.filter(test=test).count()}"
    await message.answer(msg[:4000])  

@router.message(F.chat.id == ADMIN, F.text.startswith("edit "))
async def edit_test(message: Message):
    parts = message.text.split(maxsplit=2)
    if len(parts) < 3:
        await message.answer("To'g'ri format: edit <test_id> <yangi_test_kaliti>\nMasalan: edit 123 1a2b3c10d")
        return

    _, test_code_raw, new_test_keys = parts
    try:
        test_id = int(test_code_raw)
    except ValueError:
        await message.answer("Test ID butun son bo'lishi kerak.")
        return

    try:
        user = await User.get(tg_id=message.from_user.id)
    except DoesNotExist:
        await message.answer("Foydalanuvchi topilmadi.")
        return

    test = await Tests.filter(id=test_id, user=user).first()
    if not test:
        await message.answer("Sizda bunday test mavjud emas.")
        return

    if not check_test_keys_format(new_test_keys):
        await message.answer("Yangi test kaliti noto'g'ri formatda. Iltimos, to'g'ri formatda kiriting (masalan: 1a2b3c...).")
        return

    test.test_keys = new_test_keys
    await test.save()
    await message.answer(f"Test kaliti muvaffaqiyatli yangilandi!")
    
@router.message(F.chat.id == ADMIN, F.text.startswith("hisobot "))
async def test_report(message: Message):
    parts = message.text.split()
    if len(parts) < 3:
        await message.answer("Buyruq noto'g'ri. Namuna: hisobot <test_id1> <test_id2>")
        return
    test_code1 = parts[1].strip()
    test_code2 = parts[2].strip()
    try:
        test_id1 = int(test_code1)
        test_id2 = int(test_code2)
    except ValueError:
        await message.answer("Test ID butun son bo'lishi kerak.")
        return

    try:
        user = await User.get(tg_id=message.from_user.id)
    except DoesNotExist:
        await message.answer("Foydalanuvchi topilmadi.")
        return

    tests = await Tests.filter(id__in=range(test_id1, test_id2 + 1), user=user).order_by('id').all()
    
    if not tests:
        await message.answer("Belgilangan oraliqda testlar topilmadi.")
        return
    
    await message.answer("Excel hisobot tayyorlanmoqda...")
    
    # Har bir test uchun ma'lumotlarni to'plash
    student_data = defaultdict(lambda: {'name': '', 'tests': {}, 'total_submitted': 0, 'total_correct': 0})
    
    for test in tests:
        test_answers = await UserAnswers.filter(test=test).select_related('user').order_by('-score', 'created_at').all()
        
        # Har bir test uchun maksimal savol sonini hisoblash
        test_total_questions = len(test.test_keys) // 2  # test_keys formatida har bir javob 2 ta belgi (masalan: 1a)
        
        for answer in test_answers:
            student_name = answer.user.name or f"ID:{answer.user.id}"
            student_data[answer.user.id]['name'] = student_name
            student_data[answer.user.id]['tests'][test.id] = {
                'score': answer.score,
                'time': (answer.created_at - test.created_at).total_seconds() / 60,
                'total_questions': test_total_questions
            }
            student_data[answer.user.id]['total_submitted'] += test_total_questions
            student_data[answer.user.id]['total_correct'] += answer.score
    
    # Excel fayl yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Hisoboti"
    
    # Sarlavhalar
    current_col = 1
    
    # Har bir test uchun ustun yaratish
    test_col_map = {}
    for test in tests:
        test_col_map[test.id] = current_col
        total_q = len(test.test_keys) // 2
        ws.cell(row=1, column=current_col, value=f"{test.id}({total_q})")
        ws.cell(row=1, column=current_col).font = Font(bold=True, size=12)
        ws.cell(row=1, column=current_col).alignment = Alignment(horizontal='center')
        ws.cell(row=1, column=current_col).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        current_col += 1
    
    # Har bir test uchun natijalarni yozish
    for test in tests:
        col = test_col_map[test.id]
        test_answers = await UserAnswers.filter(test=test).select_related('user').order_by('-score', 'created_at').limit(30).all()
        
        ws.cell(row=2, column=col, value="Natijalar:")
        ws.cell(row=2, column=col).font = Font(bold=True)
        
        row = 3
        for idx, answer in enumerate(test_answers, 1):
            student_name = answer.user.name or f"ID:{answer.user.id}"
            time_spent = (answer.created_at - test.created_at).total_seconds() / 60
            
            medal = ""
            if idx == 1:
                medal = "🥇 "
            elif idx == 2:
                medal = "🥈 "
            elif idx == 3:
                medal = "🥉 "
            
            text = f"{medal}{student_name} - {answer.score} ta ({round(time_spent, 1)} min)"
            ws.cell(row=row, column=col, value=text)
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
            row += 1
        
        # Jami qatnashganlar
        total_participants = await UserAnswers.filter(test=test).count()
        ws.cell(row=row, column=col, value=f"Jami qatnashganlar soni: {total_participants}")
        ws.cell(row=row, column=col).font = Font(italic=True)
    
    # Umumiy jadval yaratish
    summary_start_row = 35
    ws.cell(row=summary_start_row, column=1, value="FISH")
    ws.cell(row=summary_start_row, column=2, value=f"Yuborilgan javoblar soni({sum(len(t.test_keys)//2 for t in tests)})")
    ws.cell(row=summary_start_row, column=3, value="%")
    ws.cell(row=summary_start_row, column=4, value=f"To'g'ri javoblar soni({sum(len(t.test_keys)//2 for t in tests)})")
    ws.cell(row=summary_start_row, column=5, value="%")
    ws.cell(row=summary_start_row, column=6, value="Natija(o'rtacha)")
    
    # Sarlavhalarni formatlash
    for col in range(1, 7):
        ws.cell(row=summary_start_row, column=col).font = Font(bold=True)
        ws.cell(row=summary_start_row, column=col).alignment = Alignment(horizontal='center', wrap_text=True)
        ws.cell(row=summary_start_row, column=col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.cell(row=summary_start_row, column=col).font = Font(bold=True, color="FFFFFF")
    
    # O'quvchilarni ball bo'yicha saralash
    sorted_students = sorted(student_data.items(), key=lambda x: x[1]['total_correct'], reverse=True)
    
    summary_row = summary_start_row + 1
    total_questions_all = sum(len(t.test_keys)//2 for t in tests)
    
    for user_id, data in sorted_students:
        ws.cell(row=summary_row, column=1, value=data['name'])
        ws.cell(row=summary_row, column=2, value=data['total_submitted'])
        
        # Yuborilgan javoblar foizi
        submitted_percent = (data['total_submitted'] / total_questions_all * 100) if total_questions_all > 0 else 0
        ws.cell(row=summary_row, column=3, value=f"{int(submitted_percent)}%")
        
        ws.cell(row=summary_row, column=4, value=data['total_correct'])
        
        # To'g'ri javoblar foizi
        correct_percent = (data['total_correct'] / data['total_submitted'] * 100) if data['total_submitted'] > 0 else 0
        ws.cell(row=summary_row, column=5, value=f"{int(correct_percent)}%")
        
        # Rang berish (to'g'ri javoblar foiziga qarab)
        if correct_percent >= 80:
            color = "92D050"  # Yashil
        elif correct_percent >= 60:
            color = "FFFF00"  # Sariq
        elif correct_percent >= 40:
            color = "FFC000"  # To'q sariq
        else:
            color = "FF0000"  # Qizil
        
        ws.cell(row=summary_row, column=5).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # O'rtacha natija
        avg_result = f"{int(correct_percent)}%"
        ws.cell(row=summary_row, column=6, value=avg_result)
        
        summary_row += 1
    
    # Ustunlar kengligini o'rnatish
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20
    
    # Fayl saqlash
    filename = f"test_report_{test_id1}_{test_id2}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    
    # Faylni yuborish
    document = FSInputFile(filename)
    await message.answer_document(document, caption=f"Test hisoboti: {test_id1}-{test_id2}")
    
    # Faylni o'chirish
    os.remove(filename)
